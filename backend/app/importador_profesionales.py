"""Importa la lista de profesionales que la clínica ya tiene, en su formato.

Nadie va a retipear 40 médicos en un formulario. La lista suele existir: en un
Excel, en una planilla de Google, en un CSV exportado del sistema viejo. Este
módulo la lee tal como esté y la cruza contra el padrón del CPM.

Dos decisiones que importan:

- **Los encabezados se detectan, no se exigen.** Pedirle a una recepcionista
  que renombre columnas a `full_name,specialty,schedule` es pedirle que haga
  el trabajo del programa. Se reconocen las variantes que la gente usa de
  verdad: "Nombre", "Profesional", "Médico", "Especialidad", "Horario".
- **Nada se guarda a ciegas.** El import devuelve qué encontró y contra qué
  coincidió en el padrón; el alta la confirma una persona. Cargar 40 médicos
  de un archivo sin mirar es como se meten cuarenta errores de una sentada.
"""
import csv
import io
import logging
import re

from sqlalchemy.orm import Session

from . import registry
from .models import Doctor

logger = logging.getLogger("metabot.importador")

MAX_FILAS = 500

# Variantes de encabezado que usa la gente. Se comparan normalizadas.
COLUMNAS = {
    "name": ("nombre", "nombres", "profesional", "medico", "doctor", "apellido y nombre",
             "nombre y apellido", "apellidos y nombres", "full name", "name"),
    "specialty": ("especialidad", "especialidades", "area", "specialty", "rubro"),
    "schedule": ("horario", "horarios", "dias", "dias y horarios", "atencion",
                 "disponibilidad", "schedule"),
    "phone": ("telefono", "celular", "contacto", "phone", "movil"),
    "email": ("email", "correo", "mail", "e-mail"),
}


def _normalizar(texto: str) -> str:
    return registry.clave_de_nombre(texto).replace(" ", "") if texto else ""


def _mapear_columnas(encabezados: list[str]) -> dict[str, int]:
    """Qué columna del archivo corresponde a cada campo."""
    mapa: dict[str, int] = {}
    for i, bruto in enumerate(encabezados):
        limpio = re.sub(r"[^a-z ]", "", (bruto or "").strip().lower())
        limpio = registry.clave_de_nombre(bruto).strip()
        for campo, variantes in COLUMNAS.items():
            if campo in mapa:
                continue
            for v in variantes:
                if registry.clave_de_nombre(v) == limpio:
                    mapa[campo] = i
                    break
    return mapa


def _parece_encabezado(fila: list[str]) -> bool:
    """¿Esta fila son títulos de columna, aunque no los hayamos podido mapear?

    Si el separador se detectó mal, el encabezado llega como una sola celda
    ("Profesional;Especialidad;Horarios") y sin esto se cargaría un médico con
    ese nombre. Alcanza con que aparezca alguna palabra de título.
    """
    texto = registry.clave_de_nombre(" ".join(fila))
    palabras = set(texto.split())
    conocidas = {
        registry.clave_de_nombre(v)
        for variantes in COLUMNAS.values() for v in variantes
    }
    return any(c and c in palabras for c in conocidas)


def _filas_de_csv(contenido: bytes) -> list[list[str]]:
    texto = contenido.decode("utf-8-sig", errors="replace")
    # La gente exporta con coma o con punto y coma según la configuración
    # regional. Detectarlo evita el clásico "me quedó todo en una columna".
    #
    # No se usa csv.Sniffer solo: cuando se equivoca, el encabezado queda como
    # una única celda, no se reconoce ninguna columna y la fila de títulos
    # termina cargada como si fuera un médico. Se prueba cada separador y gana
    # el que parte la primera línea en más columnas.
    primera = texto.splitlines()[0] if texto.splitlines() else ""
    separador = max(",;\t", key=lambda s: len(next(csv.reader([primera], delimiter=s))))
    if len(next(csv.reader([primera], delimiter=separador))) == 1:
        try:
            separador = csv.Sniffer().sniff(texto[:4000], delimiters=",;\t").delimiter
        except csv.Error:
            separador = ","
    return [
        fila for fila in csv.reader(io.StringIO(texto), delimiter=separador)
        if any(fila)
    ]


def _filas_de_excel(contenido: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise ValueError(
            "Para leer .xlsx falta openpyxl en el servidor. Mientras tanto, "
            "exportá la planilla como CSV desde Excel y subila así."
        )
    libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = libro.active
    filas = []
    for fila in hoja.iter_rows(values_only=True):
        valores = ["" if c is None else str(c).strip() for c in fila]
        if any(valores):
            filas.append(valores)
    return filas


def leer(contenido: bytes, nombre_archivo: str) -> list[dict]:
    """Lee la planilla y devuelve los profesionales que encontró.

    No toca la base: solo interpreta el archivo. Lo que se guarda lo decide
    una persona después de ver esta lista.
    """
    if nombre_archivo.lower().endswith((".xlsx", ".xlsm")):
        filas = _filas_de_excel(contenido)
    else:
        filas = _filas_de_csv(contenido)
    if not filas:
        return []

    mapa = _mapear_columnas(filas[0])
    if "name" not in mapa:
        # Sin encabezados reconocibles se asume el orden más común y se avisa,
        # en vez de fallar: una planilla sin títulos igual sirve.
        cuerpo = filas[1:] if _parece_encabezado(filas[0]) else filas
        mapa = {"name": 0, "specialty": 1, "schedule": 2}
    else:
        cuerpo = filas[1:]

    def celda(fila: list[str], campo: str) -> str:
        i = mapa.get(campo)
        return (fila[i].strip() if i is not None and i < len(fila) else "")

    salida = []
    for fila in cuerpo[:MAX_FILAS]:
        nombre = celda(fila, "name")
        if not nombre or len(nombre) < 4:
            continue
        salida.append({
            "name": nombre[:200],
            "specialty": celda(fila, "specialty")[:200],
            "schedule": celda(fila, "schedule")[:100],
            "phone": celda(fila, "phone")[:50],
            "email": celda(fila, "email")[:200],
        })
    return salida


def _parecidos(clave: str, candidatos: list[dict], minimo: int = 2) -> list[dict]:
    """Deja solo los candidatos que comparten suficientes palabras con el nombre.

    Sugerir por un apellido suelto es peor que no sugerir: ofrecerle "Villalba
    Salinas, Auria Celeste" a quien escribió "Villalba Duarte, Rossana" invita
    a marcar a otra persona. Se pide coincidencia en dos palabras —salvo que el
    nombre de la planilla traiga una sola— y se ordena por cuántas comparten.
    """
    propias = set(clave.split())
    piso = min(minimo, len(propias)) or 1
    puntuados = []
    for c in candidatos:
        comunes = len(propias & set(registry.clave_de_nombre(c["full_name"]).split()))
        if comunes >= piso:
            puntuados.append((comunes, c))
    puntuados.sort(key=lambda p: (-p[0], p[1]["full_name"]))
    return [c for _, c in puntuados[:3]]


def previsualizar(db: Session, company_id: int, contenido: bytes,
                  nombre_archivo: str) -> dict:
    """Lee la planilla y la cruza contra el padrón, SIN guardar nada.

    Devuelve fila por fila qué se encontró y con qué coincidió, para que quien
    da de alta decida. Es la diferencia entre importar y importar a ciegas.
    """
    try:
        leidos = leer(contenido, nombre_archivo)
    except ValueError as exc:
        return {"error": str(exc), "filas": []}

    ya_cargados = {
        registry.clave_de_nombre(d.name)
        for d in db.query(Doctor).filter(Doctor.company_id == company_id).all()
    }

    filas = []
    for item in leidos:
        clave = registry.clave_de_nombre(item["name"])
        coincidencias = registry.buscar(db, texto=item["name"], limite=3)
        exacta = next(
            (c for c in coincidencias
             if registry.clave_de_nombre(c["full_name"]) == clave),
            None,
        )
        if not coincidencias:
            # La búsqueda con el nombre completo exige que TODAS las palabras
            # estén, y la planilla casi nunca lo trae igual que el padrón:
            # falta un segundo nombre, sobra el "Dra.", cambia el orden. Sin
            # este reintento, una fila con "Gimenez Rojas, Maria Esther" no
            # ofrecía ni una sugerencia teniendo 57 Giménez certificados
            # enfrente.
            #
            # Se afloja de a poco, de dos palabras a una: empezar por la más
            # larga sola llenaría la lista de homónimos.
            palabras = sorted(clave.split(), key=len, reverse=True)
            for cuantas in (2, 1):
                if len(palabras) < cuantas:
                    continue
                candidatos = registry.buscar(
                    db, texto=" ".join(palabras[:cuantas]), limite=8
                )
                coincidencias = _parecidos(clave, candidatos)
                if coincidencias:
                    break
        filas.append({
            **item,
            "ya_cargado": clave in ya_cargados,
            # `exacta` = el mismo nombre en el padrón. `sugerencias` = parecidos,
            # por si la planilla trae el nombre incompleto o mal escrito.
            "padron": exacta,
            "sugerencias": [c for c in coincidencias if c is not exacta][:3],
        })

    return {
        "archivo": nombre_archivo,
        "total": len(filas),
        "en_padron": sum(1 for f in filas if f["padron"]),
        "ya_cargados": sum(1 for f in filas if f["ya_cargado"]),
        "filas": filas,
    }


def confirmar(db: Session, company_id: int, seleccion: list[dict]) -> dict:
    """Da de alta lo que la persona confirmó de la previsualización."""
    creados, omitidos = [], []
    for item in seleccion:
        nombre = (item.get("name") or "").strip()
        if not nombre:
            continue
        clave = registry.clave_de_nombre(nombre)
        existe = next(
            (d for d in db.query(Doctor).filter(Doctor.company_id == company_id).all()
             if registry.clave_de_nombre(d.name) == clave),
            None,
        )
        if existe:
            omitidos.append({"name": nombre, "motivo": "ya estaba cargado"})
            continue
        doctor = Doctor(
            company_id=company_id, name=nombre[:200],
            specialty=(item.get("specialty") or "")[:200],
            schedule=(item.get("schedule") or "")[:100],
            phone=(item.get("phone") or "")[:50],
            email=(item.get("email") or "")[:200],
        )
        db.add(doctor)
        db.flush()
        resultado = registry.verificar(db, doctor)
        creados.append({
            "id": doctor.id, "name": doctor.name,
            "verification": resultado["verification"],
            "cert_number": doctor.cert_number,
        })
    db.commit()
    return {
        "creados": creados, "omitidos": omitidos,
        "verificados": sum(1 for c in creados if c["verification"] == "verified"),
        "no_figuran": sum(1 for c in creados if c["verification"] == "not_found"),
    }
